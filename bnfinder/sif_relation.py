import xlrd
from openpyxl import load_workbook
from pyExcelerator import *
import xlwt
import numpy as np
if __name__ == '__main__':

    f = open("C:/Users/Bubble/PycharmProjects/bnfinder/doc/data/DREAM3/test.sif", "r")
    rela_line=f.readlines()
    line=[]
    for item in rela_line:
        item = item.split("\t",1)
        gene1=item[0]
        item=item[1].split("\t")
        rr=item[0]
        gene2=item[1].split("\n")
        gene2 = gene2[0]

        item=[gene1,rr,gene2]
        line.append(item)

    GSElen = load_workbook(filename=r'test.xlsx')
    gseleng = GSElen.get_sheet_by_name("Sheet1")
    num = 1
    gene_name = []
    while 1:
        cell = gseleng.cell(row=num, column=1).value
        if cell:

            gn =cell.encode('unicode-escape').decode('string_escape')
            gene_name.append([num, gn])
            num = num + 1
        else:
            print(num)
            break
    num=num-1

    re = [[0 for j in range(0, num)] for i in range(0, num)]
    i = 0
    aa = 0
    ss = 0
    numa = 0
    nums = 0
    for item in line:
        i = i + 1
        print 'i', i
        regene = item[0]
        ii = item[1]
        tagene = item[2]
        if ii == '+':
            ss = ss + 1
        if ii == '-':
            aa = aa + 1
        for no in gene_name:
            if regene == no[1]:
                reno = no[0]
                break
        for no in gene_name:
            if tagene == no[1]:
                tano = no[0]
                break

        # if ii == '+':
        re[reno - 1][tano - 1] = ii
        #     nums = nums + 1
        # elif ii == '-':
        #     re[reno - 1][tano - 1] = -1
        #
        #     numa = numa + 1

    nummm = 0
    for i in range(0, num):
        for j in range(0, num):
            if re[i][j] == 1:
                nummm = nummm + 1
            if re[i][j] == -1:
                nummm = nummm + 1
    print nummm
    print re
    a0 = 0
    for i in range(0, num):
        for j in range(0, num):
            if re[i][j] == 0:
                a0 = a0 + 1
    w = Workbook()
    ws = w.add_sheet('relation')
    for i in range(0, num):
        for j in range(0, num):
            ws.write(i, j, re[i][j])
    w.save('relation.xlsx')
    aa = 1