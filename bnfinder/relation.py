import xlrd
from openpyxl import load_workbook
from pyExcelerator import *
import xlwt
import numpy as np
if __name__== '__main__':
    gene=[]
    genename_222=['BAP1','NCOR1','CDH1-2','MYC','SMARCD1','MAP2K4','APC','ZNF217','KRAS','MB','BRCA1','CDKN2A-1','AKT1','CASP8','BRCA2','CCND1-1','CDKN1B','GATA3','SF3B1','TP53','RB1','SETD2','FGFR1','ERBB2','ARID1A','TBX3','ZNF703','MDM2','ARID1B','PTEN','AKT2','ARID2','MAP3K1-2','MAP3K13-1','PIK3CA-2','SMAD4','ASXL1','STK11','COG3','RAPGEF6','SMC1A','FAM120A','ZNF740','DDX6','TPM4','MMP14','AP2B1','CALR','HNRNPA0','RAB5C','TOP2A','DARS','TUBG1','RBM6','BIRC5','FOXM1','TYMS','CCNB2','TAF15','MIR6787SLC16A3','CDC20','UBE2C','SPAG5','TRIM14','CDK1','MAD2L1','CCNA2','LSM1','PTTG1','BUB1B','DLGAP5','DMXL1','DAPK3MIR637','AURKA','NDC80','LOC101930150NF1','DHX30','KIF11','NEK2','FOXA1','MELK','EEA1','PLK4','CENPASLC35F6','RAD51','CENPE','CDC25C','ESR1','STIL','GINS1','ZC3H7B','USP34','CENPF','MKRN4PMKRN4P','RP11-403P17.4','BAX','PTTG3P','EIF4G1','TMED9','CAND1','TOP1','ILF3','KDM2A','AGR2','QRICH1','KLHDC10','SMC3','PROSC','KIF2C','ASH2L','BUB1','HMMR','CDKN3','SPC25','TPX2','MOGS','GRB7','YWHAE','CLPTM1','CFLAR','PRRC2C','WNK1','MKI67','DDHD2','ZBTB7A','SPDEF','OIP5','CEP350','RPS10','RPL37A','ATP6V0E1','CCNB1','ZNF160','APP','PRDX2','PGF','DLST','SKA1','PRC1','PPDPF','NUSAP1','MLPH','KIF4A','U2AF2','DAZAP1','CEP55','NCAPG','HJURP','KIF20A','HOOK2','BRF2','SLC35E3','KIF15','NEIL3','ASPM','HAUS2','MCM10','SLC35E1','TMEM223','FUZ','CDCA8','EIF4EBP1','ERLIN2','YIPF2','KIF18B','RACGAP1','ANLN','SLC2A4RG','DTL','TIMM22','DEPDC1','UBE2T','CDCA3','NUF2','C9orf64','WHSC1L1','TM2D2','ERGIC1','ANAPC16','CDCA5','ATAD1','CHMP4B','RAB11FIP1','PHF5A','FAM83D','PPAPDC1B','MALAT1','ZNF18','DEPDC1B','JOSD2','PARPBP','BAG4','AGR3','CASC5','ZNF397','ANKHD1','ZBTB42','RBM5','ATF7IP','PRR11','CDK13','NLN','BOD1L1','ORMDL3','PATL1','UBXN2A','TXNL1','THAP6','SYNJ2BPSYNJ2BP-COX16','LOC101929336RIF1','UACA','CCDC152','GAS2L3','OTUD4','POLI','PIK3C3','NUMBL','AFF4','LOC101060747PDPK1','SF3A2','ESPL1','PGAP3']
    genename_548=['BAP1','NCOR1','CDH1-2','MYC','SMARCD1','MAP2K4','APC','ZNF217','KRAS','MB','BRCA1','CDKN2A-1','AKT1','CASP8','BRCA2','CCND1-1','CDKN1B','GATA3','SF3B1','TP53','RB1','SETD2','FGFR1','ERBB2','ARID1A','TBX3','ZNF703','MDM2','ARID1B','PTEN','AKT2','ARID2','MAP3K1-2','MAP3K13-1','PIK3CA-2','SMAD4','ASXL1','STK11','NEXN','ANLN','PHC3','RECQL4','C5orf24','CENPL','COG3','SUV39H2','C1S','RAPGEF6','BAGE2BAGE4','SMC1A','POSTN','FAM120A','BTBD19','ITGBL1','RP11-524D16__A.3','LETM2','LUZP1','ZNF740','LOC441461','ARHGAP11BLOC100288637','CLIP1','SFXN3','MCCC2','DDX6','LOC441155ZC3H11A','TPM4','MMP14','AP2B1','LRP1','H2AFZ','CALR','ACTA2','SERPING1','EIF5B','ANP32A','HNRNPA0','MYL9','MMP2','KPNA2','RAB5C','HTRA1','PCNA','PLS3','TOP2A','VIM','DPYSL3','COL6A3','LAMB1','CALD1','DARS','HSPG2','TUBG1','AEBP1','RRM2','CKS1B','MCM6','RBM6','NID1','BIRC5','MCM2','GAS7','LAMA4','LOC101928916NNMT','PLK1','PDGFRB','G6PD','SERPINF1','PDAP1','COL1A1','TK1','SPOCK1','ADAM9','COL1A2','CTSK','KIAA0101','FOXM1','TYMS','CCNB2','FBN1','UBE2S','TAF15','MIR6787SLC16A3','CDC20','SUCLA2','ADAM12','UBE2C','FBLN1','LOXL2','TIMELESS','THBS2','PDGFRA','SPAG5','TRIM14','CDK1','EZH2','MAD2L1','PDLIM7','MFAP2','CCNA2','TMPO','LSM1','PTTG1','LOXL1','GTF2H4','ALG11UTP14C','BUB1B','DLGAP5','DMXL1','LRRC32','DAPK3MIR637','HEPH','CNN1','CDC6','ZWINT','TRIP13','TPM2','AURKA','NID2','RAD51AP1','NDC80','EMILIN1','DBF4','PKMYT1','GTSE1','COL11A1','LOC101930150NF1','DHX30','FLRT2LOC100506718','KIF11','GAS1','DZIP1','EXO1','ACP5','NEK2','FOXA1','KIF23','NAP1L3','TTK','MELK','EEA1','PLK4','SRPX','CENPASLC35F6','RAD51','CENPE','GFPT2','LAMA2','CDC25C','AP1S1','ESR1','PDGFRL','KIF20B','STIL','LRRC17','RECK','PLAU','SRPX2','RUNX1T1','TAGLN','BLM','PRKD1','OMD','COL10A1','PRRX1','TNFAIP6','ECM2','GINS1','ZC3H7B','FOXF2','LOC729966PDE4C','WISP1','NFIC','CDH11','ISLR','USP34','CENPF','MKRN4PMKRN4P','ACVR1B','RP11-403P17.4','BAX','PTTG3P','EIF4G1','TMED9','FSTL1','PTRF','HMGB2','CAND1','THY1','ATRX','TOP1','ILF3','SEPHS1','KDM2A','HNRNPDL','AGR2','QRICH1','TNPO1','KLHDC10','SMC3','PLAGL1','DCN','EFEMP2','PROSC','KIF2C','SPON1','AURKB','RARRES2','ASH2L','NDN','MXRA5','BUB1','TGFB1I1','HMMR','CDKN3','SPC25','FAP','TPX2','MOGS','GRB7','YWHAE','CLPTM1','COL3A1','CLEC11A','CFLAR','PRRC2C','WNK1','SZRD1','PXDN','MKI67','RPL27A','C1R','TM9SF4','ZC3H13','FN1','COL5A1','SPARC','DDHD2','LOC100996668ZEB1','JAM3','COL14A1','COL6A1','ANGPTL2','FANCI','PRKCDBP','HEG1','YTHDC2','OLFML2B','SNAI2','COL6A2','ZBTB7A','MXRA8','BICC1','SPDEF','IQGAP1','UBE2I','SGCD','EP300','OIP5','GGCX','MFAP5','CFH','TCF4','BGN','LRRC15','CEP350','RPS10','RPL37A','ATP6V0E1','MINK1','DKK3','PPFIBP1','CCNB1','ZNF160','CENPI','LOC101929740ZNF37A','MIR636SRSF2','APP','DNASE2','PRDX2','NFAT5','PGF','DLST','CFHCFHR1','LOX','VCAN','ANKRD12','OLFML1','SKA1','CDC27','PRC1','PPDPF','NUSAP1','ASF1B','MLPH','TACC3','KIF4A','U2AF2','DAZAP1','CMC2','CEP55','LOC100130872SPON2','NCAPG','HJURP','KIF20A','HOOK2','THOC6','CENPU','ISOC2','YEATS4','BRF2','BRF2','SLC35E3','DSCC1','ASPN','PBK','DACT1','KIF15','SYNDIG1','KDELC1','NEIL3','POLQ','CENPN','COPZ2','NOX4','ZFPM2','ECT2','ASPM','E2F8','HAUS2','ACKR4','CORIN','KERA','MCM10','MZT2B','LEPRE1','SLC35E1','TMEM223','C1QTNF3','FUZ','KIF18A','CDCA8','EIF4EBP1','CRISPLD2','ERLIN2','FAM64A','COL5A2','PDPN','COL8A2','YIPF2','RPL38','KIF18B','RACGAP1','CNOT2','THRAP3','ZFAND3','SLC2A4RG','DTL','OGN','ATAD2','TIMM22','CENPK','ITGA11','SFRP2','UBE2T','TCF19','GJB2','CDCA3','NUF2','TSHZ3','C1QTNF5MFRP','SGIP1','C9orf64','MND1','WHSC1L1','MCM8','TM2D2','LLPH','ERGIC1','ANAPC16','ANTXR1','CMTM3','CDCA5','MYLK','ATAD1','CHMP4B','RAB11FIP1','CCDC80','KNSTRN','PHF5A','MIR100HG','DDR2','UHRF1','CTHRC1','FAM83D','SLC35F6','MSRB3','FAM72AFAM72BFAM72CFAM72D','MIDN','PRICKLE1','CENPO','COL8A1','ADAMTS2','PPAPDC1B','PODN','CDCA2','CNRIP1','MIR4657PURB','ZNF18','CLMP','FNDC1','CENPW','DEPDC1B','ADAMTS12','LINC01279','GLT8D2','JOSD2','INHBA','SKA3','CNTN1','CACUL1','MAGI2-AS3','PARPBP','E2F7','MTFR2','GPX8','BAG4','AGR3','DEPDC7','CASC5','SDAD1','GALNT15','PCDH7','CYS1','ST6GAL2','ZNF397','CCDC186','ANKHD1','IQGAP3','ZNF367','LUM','ZBTB42','QSER1','SGOL2','BNC2','RBM5','FIBIN','BEND6','ALDH1L2','TIMP2','MALAT1','ATF7IP','KIAA1462','KIF21A','COL12A1','DNM3OS','PRR11','CDK13','MGC24103','KMT2C','AK021804','LTN1','GATAD2A','NLN','TMEM200A','BOD1L1','ORMDL3','PATL1','UBXN2A','TET2','DEPDC1','TXNL1','RBMS3','ESCO2','ESCO1','THAP6','SYNJ2BPSYNJ2BP-COX16','GXYLT2','HMCN1','PPAPDC1A','LOC101929336RIF1','KIF14','UACA','CCDC152','GAS2L3','OTUD4','POLI','PIK3C3','OCIAD1','SMCHD1','UQCRC2','HECTD1','CHSY3','NUMBL','RP1-78O14.1','C1QTNF6','AFF4','ZBED6','LOC101060747PDPK1','RAB11B','MRC2','SF3A2','ESPL1','ZNF500','SYDE1','COL5A3','PGAP3','RCN3','HNRNPUL2HNRNPUL2BSCL2']
    ll=548
    # for i in range(0,222):
    #     gene.append([i+1,genename_222[i]])
    for i in range(0, 548):
        gene.append([i + 1, genename_548[i]])
    rr=xlrd.open_workbook("out_gene_548_20.xlsx")
    rr1=rr.sheets()[0]
    rlen=load_workbook(filename=r'out_gene_548_20.xlsx')
    rleng = rlen.get_sheet_by_name("Sheet1")
    num1 = 1
    while 1:
        cell = rleng.cell(row=num1, column=1).value
        if cell:
            num1 = num1 + 1
        else:
            print(num1)
            break
    line=[]
    for i in range(0,num1-1):
        tt1=np.array(rr1.row_values(i))
        line.append(tt1)

    re = [[0 for j in range(1, 549)] for i in range(1, 549)]
    i = 0
    aa=0
    ss=0
    numa = 0
    nums = 0
    for item in line:
        i = i + 1
        print 'i',i
        regene = item[0].encode('unicode-escape').decode('string_escape')
        ii = item[1].encode('unicode-escape').decode('string_escape')
        tagene = item[2].encode('unicode-escape').decode('string_escape')
        if ii=='+':
            ss=ss+1
        if ii=='-':
            aa=aa+1
        for no in gene:
            if regene==no[1]:
                reno=no[0]

                break
        for no in gene:
            if tagene==no[1]:
                tano=no[0]

                break
        if ii=='+':
            re[tano-1][reno-1] = 1
            print tano,reno,re[tano - 1][reno - 1]
            nums = nums + 1
        elif ii=='-':
               re[tano-1][reno-1] = -1
               print tano,reno,re[tano - 1][reno - 1]
               numa = numa + 1

    num=0
    for i in range(0, ll):
        for j in range(0, ll):
            if re[i][j] == 1:
                num=num+1
            if re[i][j] == -1:
                num=num+1
    print num
    a0=0
    for i in range(0, ll):
      for j in range(0, ll):
          if re[i][j] == 0:
              a0=a0+1
    w=Workbook()
    ws=w.add_sheet('relation')
    for i in range(0,ll):
        for j in range(0, ll):
            ws.write(i,j,re[i][j])
    w.save('relation.xlsx')
    aa=1

    # for i in range(0,222):
    #     for j in range(0,222):
